# coding=utf-8
# python liwc.py 

# ------ LIWC ---------
import sys
import re
import csv
import operator
from turtle import distance
import xml.etree.ElementTree as ET
import json
from pyparsing import one_of
import xlsxwriter
from datetime import datetime
from openpyxl.workbook import Workbook
import openpyxl

#----- Roberta BERT sentiment
from ipaddress import v4_int_to_packed
import transformers
import torch
from transformers import RobertaTokenizer, RobertaForSequenceClassification

Sentiment = ["Negatief", "Positief", "Neutraal"]

def main(Sentiment):

    # ------ LIWC ---------
    printout = True
    workbook_name = 'D:\\I&E 2021-2022\\thesis\\LIWC\\LICW_Resultaten_'+ str(datetime.now().strftime("%H_%M_%S"))+'.xlsx'
    wb = Workbook()
    ws1 = wb.create_sheet('Sentiment')
    ws3 = wb.create_sheet('BERT')
    ws2 = wb.create_sheet('LIWC_BERT')
    page = wb.active
    inputfile = "D:\\I&E 2021-2022\\thesis\\LIWC\\TEST_Resultaten_"+ Sentiment +".csv"
    path_to_liwc_file = 'D:\\I&E 2021-2022\\thesis\\LIWC\\LIWC_Anika_non-official.txt'
    documents = []
    LiwcSentList = []
    BertSentList = []
    lines = []

    outfile_allcats = "liwc_all_cats_per_post.out"
    outfile_maincats = "liwc_main_cats_per_post.out"
    out_all = open(outfile_allcats,'w')
    out_main = open(outfile_maincats,'w')

    """ LIWC hierarchy
    I linguistic dimensions:
    1-11, no hierarchy

    II Psychological processes:
    12 Emotional processes: 13-19
    20 Cognitive processes: 21-26
    27 Senses and perceptual processes: 28-30
    31 Social processes: 32-36

    III Relativity/betrekkelijkheid
    37 Time: 38-40
    41 Space: 42-46

    IV Personal affairs
    47 Occupation: 48-50
    51 Leisure: 52-55
    56 Money
    57 Religion: 58-59
    60 Physical state: 61-65

    V Experimental dimensions
    66 Swearing
    """


    LIWC_hierarchy = dict() # key is subcat, value is main cat
    for subcat in range(1,12):
        LIWC_hierarchy[subcat] = subcat
    # the linguistic dimensions do not have a main category
    for subcat in range (12,20):
        LIWC_hierarchy[subcat] = 12
    for subcat in range (20,27):
        LIWC_hierarchy[subcat] = 20
    for subcat in range (27,31):
        LIWC_hierarchy[subcat] = 27
    for subcat in range (31,37):
        LIWC_hierarchy[subcat] = 31
    for subcat in range (37,41):
        LIWC_hierarchy[subcat] = 37
    for subcat in range (41,47):
        LIWC_hierarchy[subcat] = 41
    for subcat in range (47,51):
        LIWC_hierarchy[subcat] = 47
    for subcat in range (51,56):
        LIWC_hierarchy[subcat] = 51
    LIWC_hierarchy[56] = 56
    for subcat in range (57,60):
        LIWC_hierarchy[subcat] = 57
    for subcat in range (60,66):
        LIWC_hierarchy[subcat] = 60
    LIWC_hierarchy[66] = 66

    subcatnumbers_for_maincatnumbers = dict() # key is main cat, value is array of subcats
    all_liwc_cats = list()
    for k, v in LIWC_hierarchy.items():
        all_liwc_cats.append(k)
        if v in subcatnumbers_for_maincatnumbers:
            subcatnumbers_for_maincatnumbers[v].append(k)
        else:
            subcatnumbers_for_maincatnumbers[v] = [ k ]

    main_cats = list(subcatnumbers_for_maincatnumbers.keys())
    if printout:
        #print ("main category numbers:",subcatnumbers_for_maincatnumbers)
        print ("all cat numbers:",all_liwc_cats)
        print ("main cat numbers:",main_cats)


    def tokenize(t):
        text = t.lower()
        text = re.sub("\n"," ",text)
        text = re.sub(r'<[^>]+>',"",text) # remove all html markup
        text = re.sub('[^a-zèéeêëûüùôöòóœøîïíàáâäæãåA-Z0-9- \']', "", text)
        wrds = text.split()
        return wrds

    liwq_cats_per_word = dict()


    number_for_catname = dict() #key is catname, value is integer number
    catname_for_number = dict() #key is number, value is catname
    with open(path_to_liwc_file,'r',encoding='latin-1') as liwqfile:
        for row in liwqfile:
            row = row.rstrip()
            cat,words = row.split('#')
            words_for_cat = words.split(" ")
            #print (cat,words_for_cat)
            p = re.compile("(^[0-9]+).*")
            number = p.match(cat).group(1)

            cat = re.sub("^[0-9]+","",cat)
            number_for_catname[cat] = int(number)
            catname_for_number[int(number)] = cat

            for word in words_for_cat:
                cats_for_word = list()
                if word in liwq_cats_per_word:
                    cats_for_word = liwq_cats_per_word[word]
                cats_for_word.append(cat)
                liwq_cats_per_word[word] = cats_for_word

    liwqfile.close()

    #print (number_for_catname)
    #cats_sorted_by_number = sorted(number_for_catname.items(),key=operator.itemgetter(1))
    #print (cats_sorted_by_number)

    header_all = ["threadid","postid"]
    header_main = ["threadid","postid"]
    for catnumber in all_liwc_cats:
        header_all.append(catname_for_number[catnumber])
    out_all.write("\t".join(header_all)+"\n")
    for catnumber in main_cats:
        header_main.append(catname_for_number[catnumber])
    out_main.write("\t".join(header_main)+"\n")


    def count_words_per_liwc_cat (allcontent,catnumbers_to_count,printout=False):
        Positief = Negatief = 0
        ws1.append([" "])
        ws1.append([allcontent])
        allwords = tokenize(allcontent)
        if printout:
            print ('# of words in sample',str(len(allwords)),sep='\t')
        words_in_liwc_count = 0
        allwordsuniq = dict()
        liwqwordsuniq = dict()

        count_per_liwc_cat = dict()
        words_per_liwc_cat = dict()

        for word in allwords:
            if word in allwordsuniq:
                allwordsuniq[word] += 1
            else :
                allwordsuniq[word] = 1
            if word in liwq_cats_per_word:
                words_in_liwc_count += 1
                #print (word,liwq_cat_per_word[word])

                if word in liwqwordsuniq:
                    liwqwordsuniq[word] += 1
                else :
                    liwqwordsuniq[word] = 1

                cats_for_word = liwq_cats_per_word[word]
                for cat in cats_for_word:
                    words_for_cat = dict()
                    if cat in count_per_liwc_cat:
                        count_per_liwc_cat[cat] += 1
                        words_for_cat = words_per_liwc_cat[cat]
                    else:
                        count_per_liwc_cat[cat] = 1
                    if word in words_for_cat:
                        words_for_cat[word] += 1
                    else:
                        words_for_cat[word] = 1
                    words_per_liwc_cat[cat] = words_for_cat
        if printout:
            print ('# of unique words in sample',str(len(allwordsuniq)),sep='\t')
            print ('# of unique words in liwc',str(len(liwqwordsuniq)),sep='\t')


        if printout:
            print('\n\nLIWC counts\n')

        array_with_relcounts_per_cat = []
        for number in all_liwc_cats:
            cat = catname_for_number[number]
            if number in catnumbers_to_count:
                if cat in count_per_liwc_cat:
                    main_cat_number = LIWC_hierarchy[number]
                    freq = count_per_liwc_cat[cat]
                    relfreq = float(freq)/float(len(allwords))
                    freq_words = sorted(words_per_liwc_cat[cat].items(),key=operator.itemgetter(1),reverse=True)

                    if printout and cat == "Negemo" or cat == "anxiety" or cat == "anger" or cat == "Sadness":
                        print (main_cat_number,catname_for_number[main_cat_number],number,cat,freq,relfreq,freq_words,sep='\t')
                        ws1.append([str(cat),str(freq),str(freq_words)])
                        Negatief += freq
                    elif printout and cat == "PosEmo" or cat == "PosFeel" or cat == "Optimism":
                        ws1.append([str(cat),str(freq),str(freq_words)])
                        Positief += freq
                    array_with_relcounts_per_cat.append(relfreq)
                else:
                    array_with_relcounts_per_cat.append(0.0)
        if Positief > Negatief and abs(Positief-Negatief) > 1:
            pred = "Pos"
        elif  Positief < Negatief and abs(Positief-Negatief) > 1:
            pred = "Neg"
        else:
            pred = "Und"
        ws1.append(["Pos:", Positief,"Neg:", Negatief, "Pred:", pred])
        LiwcSentList.append(pred)
        return array_with_relcounts_per_cat





    if re.match(".*\.csv$",inputfile):
        with open(inputfile, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=',', quotechar='\"')
            for row in reader:
                print(row)
                content = row[0]
                documents.append(content)

        csvfile.close()


    allcontent = ""
    for doc in documents:
        # allcontent += doc+" "
        count_words_per_liwc_cat(doc,all_liwc_cats,True)

    print ('\n\n')
    print ("ALL")


    out_all.close()
    out_main.close()


    #----- Roberta BERT sentiment
    tokenizer = RobertaTokenizer.from_pretrained("pdelobelle/robBERT-dutch-books")
    model = RobertaForSequenceClassification.from_pretrained("pdelobelle/robBERT-dutch-books")
    v4_int_to_packed

    CompList = []

    if re.match(".*\.csv$",inputfile):
        with open(inputfile, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=',', quotechar='\"')
            print(reader)
            for row in reader:
                lines.append(row[0])
        csvfile.close()

    for i in lines:
        inputs = tokenizer(str(i), return_tensors="pt")
        outputs = model(**inputs)
        pred_logits = outputs.logits
        probs = pred_logits.softmax(dim=-1).detach().cpu().flatten().numpy().tolist()
        # Kijken naar ruwe logits of andere transformatie

    # # Geen undecided class:
        if probs[0] > probs[1]:
            BertSentList.append("Neg")
        elif probs[0] < probs[1]:
            BertSentList.append("Pos") 
        else:
            BertSentList.append("Und")       
            
    # Wel undiceded class
        # probs = pred_logits.sigmoid().detach().cpu().flatten().numpy().tolist()    
        # if pred_logits.detach().numpy()[0][0] < pred_logits.detach().numpy()[0][1] and (abs(pred_logits.detach().numpy()[0][0]) + abs(pred_logits.detach().numpy()[0][1]))/2 > 2:
        #     BertSentList.append("Pos")
        # elif  pred_logits.detach().numpy()[0][0] > pred_logits.detach().numpy()[0][1] and (abs(pred_logits.detach().numpy()[0][0]) + abs(pred_logits.detach().numpy()[0][1]))/2 > 2:
        #     BertSentList.append("Neg")
        # else:
        #     BertSentList.append("Und")

        # if probs[1] > 0.98:
        #     BertSentList.append("Pos")
        # elif probs[0] > 0.98:
        #     BertSentList.append("Neg")
        # else:
        #     BertSentList.append("Und")


        ws3.append([i])
        ws3.append(["Pos","Neg","diff"])
        ws3.append([probs[1],probs[0],abs(probs[0] - probs[1])])
        ws3.append([" "])
        CompList.append(abs(probs[0] - probs[1]))

    a = 0
    for i in CompList:
        a = a + i
        b = a/len(CompList)

    ws3.append([b])
    page.append(["Descr","LIWC","BERT" ])
    Agree = Disagree = One_Unknown = Double_Unknown = 0
    for i in range(0,len(LiwcSentList)):

        page.append([lines[i],LiwcSentList[i], BertSentList[i]])
        if LiwcSentList[i] == BertSentList[i]:
            Agree += 1
        elif (LiwcSentList[i] == "Pos" and BertSentList[i] == "Neg") or (LiwcSentList[i] == "Neg" and BertSentList[i] == "Pos") :  
            Disagree += 1        
        elif (LiwcSentList[i] == "Und" and BertSentList[i] != "Und") or (LiwcSentList[i] != "Und" and BertSentList[i] == "Und") :  
            One_Unknown += 1
        elif LiwcSentList[i] == "Und" and BertSentList[i] == "Und":
            Double_Unknown += 1
        else:
            print("er is iets fout gegaan")


    # Score voor alleen pos en neg, PosNeg = hij denkt pos maar het is neg

    page.append(["Sentiment =" + Sentiment, "TruePos", "TrueNeg", "NegPos", "PosNeg", "Und"])

    def ResultsNegPosSingle (SentList):
        TruePos = TrueNeg = PosNeg = NegPos = Und = 0
        for i in range(0,len(SentList)):

            if SentList[i] == "Pos" and BertSentList[i] == "Pos":    
                TruePos = TruePos + 1
            if SentList[i] == "Pos" and BertSentList[i] ==  "Neg":    
                PosNeg = PosNeg + 1
            if SentList[i] == "Neg" and BertSentList[i] ==  "Pos":    
                NegPos = NegPos + 1
            if SentList[i] == "Neg" and BertSentList[i] == "Neg":    
                TrueNeg = TrueNeg + 1
            if SentList[i] == "Und":
                Und = Und + 1

            # if SentList[i] == "Pos" and Sentiment == "Positief":    
            #     TruePos = TruePos + 1
            # if SentList[i] == "Pos" and Sentiment == "Negatief":    
            #     PosNeg = PosNeg + 1
            # if SentList[i] == "Neg" and Sentiment == "Positief":    
            #     NegPos = NegPos + 1
            # if SentList[i] == "Neg" and Sentiment == "Negatief":    
            #     TrueNeg = TrueNeg + 1
            # if SentList[i] == "Und":
            #     Und = Und + 1


        page.append(["Los model", TruePos, TrueNeg, NegPos, PosNeg, Und])

    def ResultsNegPosBoth (LiwcSentList,BertSentList):
        Sentlist = []
        TruePos = TrueNeg = PosNeg = NegPos = Und = 0
        for i in range(0,len(LiwcSentList)):

            if LiwcSentList[i] == BertSentList[i]:
                Sentlist.append(LiwcSentList[i])
            elif (LiwcSentList[i] == "Pos" and BertSentList[i] == "Neg") or (LiwcSentList[i] == "Neg" and BertSentList[i] == "Pos") :  
                Sentlist.append("Und")     
            elif (LiwcSentList[i] == "Und" and BertSentList[i] != "Und") or (LiwcSentList[i] != "Und" and BertSentList[i] == "Und") :  
                Sentlist.append("Und")
            elif LiwcSentList[i] == "Und" and BertSentList[i] == "Und":
                Sentlist.append("Und")

        for i in range(0,len(Sentlist)):

            if Sentlist[i] == "Pos" and Sentiment == "Positief":    
                TruePos = TruePos + 1
            if Sentlist[i] == "Pos" and Sentiment == "Negatief":    
                PosNeg = PosNeg + 1
            if Sentlist[i] == "Neg" and Sentiment == "Positief":    
                NegPos = NegPos + 1
            if Sentlist[i] == "Neg" and Sentiment == "Negatief":    
                TrueNeg = TrueNeg + 1
            if Sentlist[i] == "Und":
                Und = Und + 1
        page.append(["Beide", TruePos, TrueNeg, NegPos, PosNeg, Und])


    # def ResultsNegPosNeuSingle (SentList):
    #     TruePos = TrueNeg = PosNeg = NegPos = Und = 0
    #     for i in range(0,len(SentList)):

    #         if SentList[i] == "Pos" and Sentiment == "Positief":    
    #             TruePos = TruePos + 1
    #         if SentList[i] == "Pos" and Sentiment == "Negatief":    
    #             PosNeg = PosNeg + 1
    #         if SentList[i] == "Neg" and Sentiment == "Positief":    
    #             NegPos = NegPos + 1
    #         if SentList[i] == "Neg" and Sentiment == "Negatief":    
    #             TrueNeg = TrueNeg + 1
    #         if SentList[i] == "Und":
    #             Und = Und + 1

    #     page.append(["Los model", TruePos, TrueNeg, NegPos, PosNeg, Und])

    # ResultsNegPosSingle(BertSentList)
    ResultsNegPosSingle(LiwcSentList)
    # ResultsNegPosBoth(LiwcSentList,BertSentList)





    ws2.append(["Aantal geanalyseerd:", Agree+Disagree+One_Unknown+Double_Unknown])
    ws2.append(["Overeenkomst:", Agree])
    ws2.append(["Tegenspraak:", Disagree])
    ws2.append(["Een undecided:", One_Unknown])
    ws2.append(["Twee undecided:", Double_Unknown])





    wb.save(filename=workbook_name)

for i in range(0, len(Sentiment)):
    main(Sentiment[i])