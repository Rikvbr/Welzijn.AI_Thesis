    # ------ LIWC ---------
from inspect import FullArgSpec
import sys
import re
import csv
import operator
from turtle import distance
import xml.etree.ElementTree as ET
import json
from pyparsing import one_of
from sklearn.metrics import f1_score
import xlsxwriter
from datetime import datetime
from openpyxl.workbook import Workbook
import openpyxl
import random
from random import sample
#----- Roberta BERT sentiment
from ipaddress import v4_int_to_packed
import transformers
import torch
from transformers import RobertaTokenizer, RobertaForSequenceClassification
printout = True
workbook_name = 'D:\\I&E 2021-2022\\thesis\\LIWC\\Resultaten_Iteraties'+ str(datetime.now().strftime("%H_%M_%S"))+'.xlsx'
wb = Workbook()
ws1 = wb.create_sheet('Resultaten')
ws2 = wb.create_sheet('F1_scores')
page = wb.active
inputfile = "D:\\I&E 2021-2022\\thesis\\Thesis_69_respondenten.csv"
path_to_liwc_file = 'D:\\I&E 2021-2022\\thesis\\LIWC\\LIWC_Anika_non-official.txt'

lines = []
Results_total = []


thresholds = [0.999]
outfile_allcats = "liwc_all_cats_per_post.out"
outfile_maincats = "liwc_main_cats_per_post.out"
out_all = open(outfile_allcats,'w')
out_main = open(outfile_maincats,'w')

if re.match(".*\.csv$",inputfile):
    with open(inputfile, 'r', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', quotechar='\"')
        for row in reader:
            lines.append(row[0])
    csvfile.close()
    
#----- Roberta BERT sentiment
tokenizer = RobertaTokenizer.from_pretrained("pdelobelle/robBERT-dutch-books")
model = RobertaForSequenceClassification.from_pretrained("pdelobelle/robBERT-dutch-books")
v4_int_to_packed
for i in range(0,11):
    fullSampleList = []
    samplelist = []
    samplelist.append(sample(range(0, int(len(lines)/3 + 1)), k=int(len(lines)/6)))
    samplelist.append(sample(range(int(len(lines)/3), int((2 *len(lines)/3)+1)), k=int(len(lines)/6)))
    samplelist.append(sample(range(int(2 *len(lines)/3), int(len(lines))), k=int(len(lines)/6)))
    for a in samplelist:
        for x in a:
            fullSampleList.append(x)

    thresholds_total = []
    for y in thresholds:

        threshold_single = []
        count = 0
        for x in range(0, len(fullSampleList)):
            local_pred = []
            if count < len(fullSampleList)/3:
                local_pred.append("Pos")
            elif count >= len(fullSampleList)/3 and count < 2*len(fullSampleList)/3:
                local_pred.append("Neg")
            else:
                local_pred.append("Und")
            count += 1

            inputs = tokenizer(str(lines[fullSampleList[x]]), return_tensors="pt")
            outputs = model(**inputs)
            pred_logits = outputs.logits
            probs = pred_logits.softmax(dim=-1).detach().cpu().flatten().numpy().tolist()   
            if probs[1] > y:
                local_pred.append("Pos")
            elif probs[0] > y:
                local_pred.append("Neg")
            else:
                local_pred.append("Und")

            threshold_single.append(local_pred)
            # print("volgende rij voor threshold")
        thresholds_total.append(threshold_single)
        print("volgende threshold: " + str(y))
    Results_total.append(thresholds_total)
    print("__volgende iteratie " + str(i))


Scores_totaal = []
for iteraties in Results_total:
    resultaten_threshold = []
    F1_local = []
    for thresholds in iteraties:
        PosPos = PosUnd = PosNeg = UndPos = UndUnd =UndNeg = NegPos =  NegUnd = NegNeg = 0
        for rijen in thresholds:
            if rijen[1] == "Pos" and rijen[0] == "Pos":    
                PosPos = PosPos + 1
            if rijen[1] == "Pos" and rijen[0] == "Und":    
                PosUnd = PosUnd + 1
            if rijen[1] == "Pos" and rijen[0] == "Neg":    
                PosNeg = PosNeg + 1
            if rijen[1] == "Und" and rijen[0] == "Pos":    
                UndPos = UndPos + 1
            if rijen[1] == "Und" and rijen[0] == "Und":    
                UndUnd = UndUnd + 1
            if rijen[1] == "Und" and rijen[0] == "Neg":    
                UndNeg = UndNeg + 1
            if rijen[1] == "Neg" and rijen[0] == "Pos":    
                NegPos = NegPos + 1
            if rijen[1] == "Neg" and rijen[0] == "Und":    
                NegUnd = NegUnd + 1
            if rijen[1] == "Neg" and rijen[0] == "Neg":    
                NegNeg = NegNeg + 1
        resultaten_threshold.append([PosPos, PosUnd, PosNeg, UndPos, UndUnd, UndNeg, NegPos, NegUnd, NegNeg])
        
        try:
            precision_pos = PosPos / (PosPos + PosUnd + PosNeg)
            precision_neu = UndUnd / (UndUnd + UndPos + UndNeg)
            precision_neg = NegNeg / (NegNeg + NegUnd + NegPos)

            Recall_pos = PosPos /(PosPos + UndPos + NegPos)
            Recall_neu = UndUnd /(UndUnd + NegUnd + PosUnd)
            Recall_neg = NegNeg /(NegNeg + UndNeg + PosNeg)

            f1_pos = 2* (precision_pos * Recall_pos)/( precision_pos + Recall_pos)
            f1_neu = 2* (precision_neu * Recall_neu)/( precision_neu + Recall_neu)
            f1_neg = 2* (precision_neg * Recall_neg)/( precision_neg + Recall_neg)

            average_f1 = (f1_pos + f1_neu + f1_neg) / 3

            ws1.append(["","Pos", "Neu", "Neg"])
            ws1.append(["Pos",PosPos, PosUnd, PosNeg])
            ws1.append(["Neu",UndPos, UndUnd, UndNeg])
            ws1.append(["Neg",NegPos, NegUnd, NegNeg])
            ws1.append([" "])
            ws1.append(["Recall_pos", Recall_pos, " ", "Precision_pos", precision_pos ])
            ws1.append(["Recall_neu", Recall_neu, " ", "Precision_neu", precision_neu ])
            ws1.append(["Recall_neg", Recall_neg, " ", "Precision_neg", precision_neg ])
            ws1.append([" "])
            ws1.append(["F1", average_f1])
            ws1.append([" "])
            F1_local.append(average_f1)
        except:
            ws1.append(["Pos",PosPos, PosUnd, PosNeg])
            ws1.append(["Neu",UndPos, UndUnd, UndNeg])
            ws1.append(["Neg",NegPos, NegUnd, NegNeg])
            ws1.append([" "])
            F1_local.append(0)
        Scores_totaal.append(resultaten_threshold)
        ws1.append([" "])
    ws2.append(F1_local)
print(Scores_totaal)
wb.save(filename=workbook_name)        
