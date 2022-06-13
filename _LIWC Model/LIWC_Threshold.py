    # ------ LIWC ---------
from inspect import FullArgSpec
import sys
import re
import csv
import operator
from turtle import distance
import xml.etree.ElementTree as ET
import json
from pyparsing import one_of
import xlsxwriter
from datetime import datetime
from openpyxl.workbook import Workbook
import openpyxl
import random
from random import sample
#----- Roberta BERT sentiment
from ipaddress import v4_int_to_packed
import transformers
import torch
from transformers import RobertaTokenizer, RobertaForSequenceClassification
printout = False
workbook_name = 'D:\\I&E 2021-2022\\thesis\\LIWC\\Resultaten_Iteraties'+ str(datetime.now().strftime("%H_%M_%S"))+'.xlsx'
wb = Workbook()
ws1 = wb.create_sheet('Resultaten')
ws2 = wb.create_sheet('F1_scores')
page = wb.active
inputfile = "D:\\I&E 2021-2022\\thesis\\Thesis_respondenten.csv"
path_to_liwc_file = 'D:\\I&E 2021-2022\\thesis\\LIWC\\LIWC_Anika_non-official.txt'
documents = []
lines = []
Results_total = []


thresholds = [0,1,2,3]
outfile_allcats = "liwc_all_cats_per_post.out"
outfile_maincats = "liwc_main_cats_per_post.out"
out_all = open(outfile_allcats,'w')
out_main = open(outfile_maincats,'w')

if re.match(".*\.csv$",inputfile):
    with open(inputfile, 'r', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', quotechar='\"')
        for row in reader:
            lines.append(row[0])
            content = row[0]
            documents.append(content)

    csvfile.close()

    
LIWC_hierarchy = dict() # key is subcat, value is main cat
for subcat in range(1,12):
    LIWC_hierarchy[subcat] = subcat
# the linguistic dimensions do not have a main category
for subcat in range (12,20):
    LIWC_hierarchy[subcat] = 12
for subcat in range (20,27):
    LIWC_hierarchy[subcat] = 20
for subcat in range (27,31):
    LIWC_hierarchy[subcat] = 27
for subcat in range (31,37):
    LIWC_hierarchy[subcat] = 31
for subcat in range (37,41):
    LIWC_hierarchy[subcat] = 37
for subcat in range (41,47):
    LIWC_hierarchy[subcat] = 41
for subcat in range (47,51):
    LIWC_hierarchy[subcat] = 47
for subcat in range (51,56):
    LIWC_hierarchy[subcat] = 51
LIWC_hierarchy[56] = 56
for subcat in range (57,60):
    LIWC_hierarchy[subcat] = 57
for subcat in range (60,66):
    LIWC_hierarchy[subcat] = 60
LIWC_hierarchy[66] = 66

subcatnumbers_for_maincatnumbers = dict() # key is main cat, value is array of subcats
all_liwc_cats = list()
for k, v in LIWC_hierarchy.items():
    all_liwc_cats.append(k)
    if v in subcatnumbers_for_maincatnumbers:
        subcatnumbers_for_maincatnumbers[v].append(k)
    else:
        subcatnumbers_for_maincatnumbers[v] = [ k ]

main_cats = list(subcatnumbers_for_maincatnumbers.keys())
if printout:
    #print ("main category numbers:",subcatnumbers_for_maincatnumbers)
    print ("all cat numbers:",all_liwc_cats)
    print ("main cat numbers:",main_cats)


def tokenize(t):
    text = t.lower()
    text = re.sub("\n"," ",text)
    text = re.sub(r'<[^>]+>',"",text) # remove all html markup
    text = re.sub('[^a-zèéeêëûüùôöòóœøîïíàáâäæãåA-Z0-9- \']', "", text)
    wrds = text.split()
    return wrds

liwq_cats_per_word = dict()


number_for_catname = dict() #key is catname, value is integer number
catname_for_number = dict() #key is number, value is catname
with open(path_to_liwc_file,'r',encoding='latin-1') as liwqfile:
    for row in liwqfile:
        row = row.rstrip()
        cat,words = row.split('#')
        words_for_cat = words.split(" ")
        #print (cat,words_for_cat)
        p = re.compile("(^[0-9]+).*")
        number = p.match(cat).group(1)

        cat = re.sub("^[0-9]+","",cat)
        number_for_catname[cat] = int(number)
        catname_for_number[int(number)] = cat

        for word in words_for_cat:
            cats_for_word = list()
            if word in liwq_cats_per_word:
                cats_for_word = liwq_cats_per_word[word]
            cats_for_word.append(cat)
            liwq_cats_per_word[word] = cats_for_word

liwqfile.close()
header_all = ["threadid","postid"]
header_main = ["threadid","postid"]
for catnumber in all_liwc_cats:
    header_all.append(catname_for_number[catnumber])
out_all.write("\t".join(header_all)+"\n")
for catnumber in main_cats:
    header_main.append(catname_for_number[catnumber])
out_main.write("\t".join(header_main)+"\n")


def count_words_per_liwc_cat (allcontent,catnumbers_to_count, threshold ,printout):
        Positief = Negatief = 0
        # ws1.append([" "])
        # ws1.append([allcontent])
        allwords = tokenize(allcontent)
        if printout:
            print ('# of words in sample',str(len(allwords)),sep='\t')
        words_in_liwc_count = 0
        allwordsuniq = dict()
        liwqwordsuniq = dict()

        count_per_liwc_cat = dict()
        words_per_liwc_cat = dict()

        for word in allwords:
            if word in allwordsuniq:
                allwordsuniq[word] += 1
            else :
                allwordsuniq[word] = 1
            if word in liwq_cats_per_word:
                words_in_liwc_count += 1
                #print (word,liwq_cat_per_word[word])

                if word in liwqwordsuniq:
                    liwqwordsuniq[word] += 1
                else :
                    liwqwordsuniq[word] = 1

                cats_for_word = liwq_cats_per_word[word]
                for cat in cats_for_word:
                    words_for_cat = dict()
                    if cat in count_per_liwc_cat:
                        count_per_liwc_cat[cat] += 1
                        words_for_cat = words_per_liwc_cat[cat]
                    else:
                        count_per_liwc_cat[cat] = 1
                    if word in words_for_cat:
                        words_for_cat[word] += 1
                    else:
                        words_for_cat[word] = 1
                    words_per_liwc_cat[cat] = words_for_cat
        if printout:
            print ('# of unique words in sample',str(len(allwordsuniq)),sep='\t')
            print ('# of unique words in liwc',str(len(liwqwordsuniq)),sep='\t')


        if printout:
            print('\n\nLIWC counts\n')

        array_with_relcounts_per_cat = []
        for number in all_liwc_cats:
            cat = catname_for_number[number]
            if number in catnumbers_to_count:
                if cat in count_per_liwc_cat:
                    main_cat_number = LIWC_hierarchy[number]
                    freq = count_per_liwc_cat[cat]
                    relfreq = float(freq)/float(len(allwords))
                    freq_words = sorted(words_per_liwc_cat[cat].items(),key=operator.itemgetter(1),reverse=True)

                    if cat == "Negemo" or cat == "anxiety" or cat == "anger" or cat == "Sadness":
                        # print (main_cat_number,catname_for_number[main_cat_number],number,cat,freq,relfreq,freq_words,sep='\t')
                        # ws1.append([str(cat),str(freq),str(freq_words)])
                        Negatief += freq
                    elif cat == "PosEmo" or cat == "PosFeel" or cat == "Optimism":
                        # ws1.append([str(cat),str(freq),str(freq_words)])
                        Positief += freq
                    array_with_relcounts_per_cat.append(relfreq)
                else:
                    array_with_relcounts_per_cat.append(0.0)
        if Positief > Negatief and abs(Positief-Negatief) > threshold:
            pred = "Pos"
        elif  Positief < Negatief and abs(Positief-Negatief) > threshold:
            pred = "Neg"
        else:
            pred = "Und"
        # ws1.append(["Pos:", Positief,"Neg:", Negatief, "Pred:", pred])
        
        return pred

v4_int_to_packed
for i in range(0,11):
    fullSampleList = []
    samplelist = []
    samplelist.append(sample(range(0, int(len(lines)/3 + 1)), k=int(len(lines)/6)))
    samplelist.append(sample(range(int(len(lines)/3), int((2 *len(lines)/3)+1)), k=int(len(lines)/6)))
    samplelist.append(sample(range(int(2 *len(lines)/3), int(len(lines))), k=int(len(lines)/6)))
    for a in samplelist:
        for x in a:
            fullSampleList.append(x)
    thresholds_total = []
    for y in thresholds:
        count = 0
        threshold_single = []
        for x in range(0, len(fullSampleList)):
            local_pred = []
            if count < len(fullSampleList)/3:
                local_pred.append("Pos")
            elif count >= len(fullSampleList)/3 and count < 2*len(fullSampleList)/3:
                local_pred.append("Neg")
            else:
                local_pred.append("Und")
            count += 1
            resultaat = count_words_per_liwc_cat(documents[fullSampleList[x]],all_liwc_cats, y, False)

            local_pred.append(resultaat)
            threshold_single.append(local_pred)
        thresholds_total.append(threshold_single)
        print("volgende threshold: " + str(y))
    Results_total.append(thresholds_total)
    print("__volgende iteratie " + str(i))


Scores_totaal = []
for iteraties in Results_total:
    resultaten_threshold = []
    F1_local = []
    Precision_local= []
    Recall_local= []
    for thresholds in iteraties:
        PosPos = PosUnd = PosNeg = UndPos = UndUnd =UndNeg = NegPos =  NegUnd = NegNeg = 0
        for rijen in thresholds:
            if rijen[1] == "Pos" and rijen[0] == "Pos":    
                PosPos = PosPos + 1
            if rijen[1] == "Pos" and rijen[0] == "Und":    
                PosUnd = PosUnd + 1
            if rijen[1] == "Pos" and rijen[0] == "Neg":    
                PosNeg = PosNeg + 1
            if rijen[1] == "Und" and rijen[0] == "Pos":    
                UndPos = UndPos + 1
            if rijen[1] == "Und" and rijen[0] == "Und":    
                UndUnd = UndUnd + 1
            if rijen[1] == "Und" and rijen[0] == "Neg":    
                UndNeg = UndNeg + 1
            if rijen[1] == "Neg" and rijen[0] == "Pos":    
                NegPos = NegPos + 1
            if rijen[1] == "Neg" and rijen[0] == "Und":    
                NegUnd = NegUnd + 1
            if rijen[1] == "Neg" and rijen[0] == "Neg":    
                NegNeg = NegNeg + 1
        resultaten_threshold.append([PosPos, PosUnd, PosNeg, UndPos, UndUnd, UndNeg, NegPos, NegUnd, NegNeg])
        try:
            precision_pos = PosPos / (PosPos + PosUnd + PosNeg)
            precision_neu = UndUnd / (UndUnd + UndPos + UndNeg)
            precision_neg = NegNeg / (NegNeg + NegUnd + NegPos)

            Recall_pos = PosPos /(PosPos + UndPos + NegPos)
            Recall_neu = UndUnd /(UndUnd + NegUnd + PosUnd)
            Recall_neg = NegNeg /(NegNeg + UndNeg + PosNeg)

            f1_pos = 2* (precision_pos * Recall_pos)/( precision_pos + Recall_pos)
            f1_neu = 2* (precision_neu * Recall_neu)/( precision_neu + Recall_neu)
            f1_neg = 2* (precision_neg * Recall_neg)/( precision_neg + Recall_neg)

            average_f1 = (f1_pos + f1_neu + f1_neg) / 3

            ws1.append(["","Pos", "Neu", "Neg"])
            ws1.append(["Pos",PosPos, PosUnd, PosNeg])
            ws1.append(["Neu",UndPos, UndUnd, UndNeg])
            ws1.append(["Neg",NegPos, NegUnd, NegNeg])
            ws1.append([" "])
            ws1.append(["Recall_pos", Recall_pos, " ", "Precision_pos", precision_pos ])
            ws1.append(["Recall_neu", Recall_neu, " ", "Precision_neu", precision_neu ])
            ws1.append(["Recall_neg", Recall_neg, " ", "Precision_neg", precision_neg ])
            ws1.append([" "])
            ws1.append(["F1", average_f1])
            ws1.append([" "])
            F1_local.append(average_f1)

        except:
            ws1.append(["Pos",PosPos, PosUnd, PosNeg])
            ws1.append(["Neu",UndPos, UndUnd, UndNeg])
            ws1.append(["Neg",NegPos, NegUnd, NegNeg])
            ws1.append([" "])
            F1_local.append(0)
        Scores_totaal.append(resultaten_threshold)
        ws1.append([" "])
    ws2.append(F1_local)
print(Scores_totaal)
wb.save(filename=workbook_name)        
